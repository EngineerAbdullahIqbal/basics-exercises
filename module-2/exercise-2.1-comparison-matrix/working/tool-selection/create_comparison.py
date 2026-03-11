#!/usr/bin/env python3
"""Create tool comparison spreadsheet with weighted scoring matrix."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Tool Comparison Matrix"

# Data based on research for Trello, Asana, Notion, Linear
# Scores are 1-5 (5 = best)
# Weights: High = 3, Medium = 2, Low = 1

criteria_data = [
    # (Criterion, Weight, Weight Value, Trello, Asana, Notion, Linear)
    ("Price per user (paid tier)", "High", 3, 4, 3, 4, 3),
    ("Learning curve for non-technical users", "High", 3, 5, 4, 3, 2),
    ("Slack + Google Workspace integration quality", "High", 3, 4, 5, 4, 4),
    ("Mobile app quality", "Medium", 2, 4, 5, 3, 3),
    ("Reporting / workload visibility", "Medium", 2, 3, 5, 4, 5),
    ("Scalability to 25 users", "Medium", 2, 4, 5, 5, 4),
    ("Free tier usefulness", "Low", 1, 4, 3, 5, 2),
]

tools = ["Trello", "Asana", "Notion", "Linear"]

# Styling
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=12, color="FFFFFF")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')

# Color fills for decision matrix
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 4-5
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 3
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 1-2

# Write headers
ws.cell(row=1, column=1, value="Criteria").font = header_font_white
ws.cell(row=1, column=2, value="Weight").font = header_font_white
ws.cell(row=1, column=3, value="Weight Value").font = header_font_white

for i, tool in enumerate(tools, start=4):
    ws.cell(row=1, column=i, value=tool).font = header_font_white

for col in range(1, 8):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# Write criteria and scores
for row_idx, (criterion, weight, weight_val, trello, asana, notion, linear) in enumerate(criteria_data, start=2):
    ws.cell(row=row_idx, column=1, value=criterion).border = border
    ws.cell(row=row_idx, column=2, value=weight).border = border
    ws.cell(row=row_idx, column=3, value=weight_val).border = border
    ws.cell(row=row_idx, column=4, value=trello).border = border
    ws.cell(row=row_idx, column=5, value=asana).border = border
    ws.cell(row=row_idx, column=6, value=notion).border = border
    ws.cell(row=row_idx, column=7, value=linear).border = border

# Calculate weighted scores row
row_idx = len(criteria_data) + 2
ws.cell(row=row_idx, column=1, value="Weighted Total Score").font = Font(bold=True)
for col in range(1, 4):
    ws.cell(row=row_idx, column=col).border = border

# Calculate weighted totals for each tool
weighted_scores = []
for tool_idx, tool_scores in enumerate([(4, 5, 4, 4, 3, 4, 4),  # Trello
                                         (3, 4, 5, 5, 5, 5, 3),  # Asana
                                         (4, 3, 4, 3, 4, 5, 5),  # Notion
                                         (3, 2, 4, 3, 5, 4, 2)], start=0):  # Linear
    weights = [3, 3, 3, 2, 2, 2, 1]
    weighted_total = sum(s * w for s, w in zip(tool_scores, weights))
    weighted_scores.append(weighted_total)
    ws.cell(row=row_idx, column=4 + tool_idx, value=weighted_total).font = Font(bold=True, size=12)
    ws.cell(row=row_idx, column=4 + tool_idx).border = border

# Column widths
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12

# Create Decision Matrix sheet (color-coded)
ws2 = wb.create_sheet(title="Decision Matrix")

# Copy headers
for col in range(1, 8):
    cell = ws2.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border

# Copy criteria data with color coding
for row_idx in range(2, len(criteria_data) + 2):
    criterion = ws.cell(row=row_idx, column=1).value
    weight = ws.cell(row=row_idx, column=2).value
    weight_val = ws.cell(row=row_idx, column=3).value
    
    ws2.cell(row=row_idx, column=1, value=criterion).border = border
    ws2.cell(row=row_idx, column=2, value=weight).border = border
    ws2.cell(row=row_idx, column=3, value=weight_val).border = border
    
    for col_idx, tool_idx in enumerate(range(4, 8), start=1):
        score = ws.cell(row=row_idx, column=tool_idx).value
        cell = ws2.cell(row=row_idx, column=col_idx + 3, value=score)
        cell.border = border
        
        # Apply color coding
        if score >= 4:
            cell.fill = green_fill
        elif score == 3:
            cell.fill = yellow_fill
        else:
            cell.fill = red_fill

# Add weighted totals row
for col in range(1, 8):
    cell = ws2.cell(row=row_idx + 2, column=col, value=ws.cell(row=len(criteria_data) + 2, column=col).value)
    cell.border = border
    if col >= 4:
        cell.font = Font(bold=True, size=12)

# Add color legend
legend_row = row_idx + 4
ws2.cell(row=legend_row, column=1, value="Color Legend:").font = Font(bold=True)
ws2.cell(row=legend_row + 1, column=2, value="Green (4-5)").fill = green_fill
ws2.cell(row=legend_row + 1, column=2, value="Green (4-5) = Excellent").border = border
ws2.cell(row=legend_row + 2, column=2, value="Yellow (3)").fill = yellow_fill
ws2.cell(row=legend_row + 2, column=2, value="Yellow (3) = Good").border = border
ws2.cell(row=legend_row + 3, column=2, value="Red (1-2)").fill = red_fill
ws2.cell(row=legend_row + 3, column=2, value="Red (1-2) = Poor").border = border

# Add summary section
summary_row = legend_row + 6
ws2.cell(row=summary_row, column=1, value="Summary Rankings:").font = Font(bold=True)
rankings = sorted(zip(tools, weighted_scores), key=lambda x: x[1], reverse=True)
for i, (tool, score) in enumerate(rankings, start=1):
    ws2.cell(row=summary_row + i, column=1, value=f"#{i}: {tool} ({score} pts)")

# Save workbook
output_path = "/home/abdullahiqbal/Abdullah/Digita-FTEs/basics-exercises/module-2/exercise-2.1-comparison-matrix/working/tool-selection/tool-comparison.xlsx"
wb.save(output_path)
print(f"Spreadsheet created: {output_path}")
